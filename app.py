import warnings
warnings.filterwarnings("ignore")

import copy
import logging
import random
import json
import yaml
import time
import pickle
import io
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Tuple, Any, Union
from dataclasses import dataclass, field, asdict

import numpy as np
import pandas as pd
import yfinance as yf
import torch
import torch.nn as nn
import torch.nn.functional as F
from torch.utils.data import DataLoader, TensorDataset, Dataset
from sklearn.linear_model import LogisticRegression
from sklearn.model_selection import TimeSeriesSplit
from sklearn.preprocessing import StandardScaler
from sklearn.feature_selection import mutual_info_classif
from hmmlearn.hmm import GaussianHMM
from scipy import stats as scipy_stats
from scipy.linalg import solve_triangular, cholesky

import optuna
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import LineChart, Reference
import streamlit as st

optuna.logging.set_verbosity(optuna.logging.WARNING)

st.set_page_config(
    page_title="Hybrid Quant Backtester",
    page_icon="📈",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ===========================================================================
# CONFIGURATION SYSTEM
# ===========================================================================

@dataclass
class DataConfig:
    ticker: str = "SPY"
    period: str = "10y"
    interval: str = "1d"

@dataclass
class BacktestConfig:
    n_splits: int = 5
    anchored: bool = False
    seed: int = 42
    n_workers: int = 1

@dataclass
class ModelConfig:
    n_states: int = 3
    architecture: str = "mlp"  
    hidden_dim: int = 64
    dropout: float = 0.25
    lr: float = 1e-3
    num_layers: int = 2
    use_student_t_hmm: bool = False
    seq_len: int = 20  
    use_mixed_precision: bool = False  

@dataclass
class TradingConfig:
    prob_long: float = 0.52
    prob_short: float = 0.48
    regime_gate: float = 0.45
    vol_target: float = 0.15
    cost_bps: float = 2.0
    spread_bps: float = 1.0  
    impact_factor: float = 0.1  
    max_churn: int = 6
    churn_window: int = 20

@dataclass
class RiskConfig:
    enabled: bool = False
    stop_loss_pct: float = 0.02
    take_profit_pct: float = 0.05
    trailing_stop_pct: float = 0.015
    max_drawdown_halt: float = 0.10
    use_conservative_gaps: bool = True  
    regime_stop_multipliers: Dict[int, float] = field(default_factory=dict)

@dataclass
class TuningConfig:
    enabled: bool = True
    n_trials: int = 15
    inner_splits: int = 3

@dataclass
class FeatureConfig:
    use_advanced_features: bool = True
    feature_selection: bool = False
    n_selected_features: int = 12

@dataclass
class OutputConfig:
    dir: str = "output"
    prefix: Optional[str] = None
    save_logs: bool = True

@dataclass
class Config:
    data: DataConfig = field(default_factory=DataConfig)
    backtest: BacktestConfig = field(default_factory=BacktestConfig)
    model: ModelConfig = field(default_factory=ModelConfig)
    trading: TradingConfig = field(default_factory=TradingConfig)
    risk: RiskConfig = field(default_factory=RiskConfig)
    tuning: TuningConfig = field(default_factory=TuningConfig)
    features: FeatureConfig = field(default_factory=FeatureConfig)
    output: OutputConfig = field(default_factory=OutputConfig)

def set_seed(seed: int):
    """Set random seeds for full reproducibility."""
    random.seed(seed)
    np.random.seed(seed)
    torch.manual_seed(seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed_all(seed)

# ===========================================================================
# DATA LAYER (Robust Validation & Retry Logic)
# ===========================================================================

@st.cache_data(show_spinner=False, ttl=3600)
def load_data(ticker: str = "SPY", period: str = "10y", interval: str = "1d", max_retries: int = 3) -> pd.DataFrame:
    """Download OHLCV data with retry logic and strict integrity validation."""
    df = pd.DataFrame()
    for attempt in range(max_retries):
        try:
            df = yf.download(ticker, period=period, interval=interval, auto_adjust=True, progress=False)
            if not df.empty:
                break
        except Exception as e:
            time.sleep(2 ** attempt)  
            
    if df.empty:
        raise ValueError(f"CRITICAL: No data returned for ticker {ticker} after {max_retries} attempts.")
    
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = [c[0].lower() for c in df.columns]
    else:
        df.columns = [str(c).lower() for c in df.columns]
    
    required_cols = {'close', 'high', 'low', 'open'}
    missing_cols = required_cols - set(df.columns)
    if missing_cols:
        raise ValueError(f"CRITICAL: Missing required OHLC data columns for {ticker}: {missing_cols}")
        
    if len(df) < 60:
        raise ValueError(f"CRITICAL: Insufficient data history for {ticker}. Received {len(df)} rows, minimum required is 60.")
        
    if (df[['close', 'high', 'low', 'open']] <= 0).any().any():
        df = df.replace(0, np.nan).ffill()
        
    return df.dropna()

# ===========================================================================
# FEATURE ENGINEERING - No Look-Ahead Bias
# ===========================================================================

def build_features(df: pd.DataFrame, use_advanced: bool = True) -> Tuple[pd.DataFrame, pd.Series, pd.Series, pd.DataFrame]:
    close = df["close"].copy()
    high = df["high"].copy() if "high" in df.columns else close.copy()
    low = df["low"].copy() if "low" in df.columns else close.copy()
    open_price = df["open"].copy() if "open" in df.columns else close.copy()
    volume = df["volume"].copy() if "volume" in df.columns else pd.Series(1.0, index=df.index, dtype=float)
    eps = 1e-8
    
    features = {}
    features["ret_1"] = np.log(close / close.shift(1))
    features["ret_5"] = np.log(close / close.shift(5))
    features["ret_20"] = np.log(close / close.shift(20))
    features["vol_20"] = features["ret_1"].rolling(20).std()
    features["vol_20_lagged"] = features["vol_20"].shift(1)  
    
    sma_10 = close.rolling(10).mean()
    sma_50 = close.rolling(50).mean()
    features["dist_sma10"] = close / sma_10 - 1.0
    features["dist_sma50"] = close / sma_50 - 1.0
    
    delta = close.diff()
    gain = delta.clip(lower=0).rolling(14).mean()
    loss = (-delta.clip(upper=0)).rolling(14).mean()
    rs = gain / (loss + eps)
    features["rsi14"] = 100 - 100 / (1 + rs)
    features["vol_chg_5"] = np.log(volume / (volume.shift(5) + eps))
    
    if use_advanced:
        ema_12 = close.ewm(span=12, adjust=False).mean()
        ema_26 = close.ewm(span=26, adjust=False).mean()
        macd = ema_12 - ema_26
        macd_signal = macd.ewm(span=9, adjust=False).mean()
        features["macd_hist"] = (macd - macd_signal) / (close + eps)
        features["roc_10"] = (close - close.shift(10)) / (close.shift(10) + eps)
        momentum_10 = close - close.shift(10)
        features["momentum_10_norm"] = momentum_10 / (close.rolling(20).std() + eps)
        bb_mid = close.rolling(20).mean()
        bb_std = close.rolling(20).std()
        features["bb_position"] = (close - bb_mid) / (2 * bb_std + eps)
        features["bb_width"] = (4 * bb_std) / (bb_mid + eps)
        tr = pd.concat([high - low, (high - close.shift(1)).abs(), (low - close.shift(1)).abs()], axis=1).max(axis=1)
        features["atr_14"] = tr.rolling(14).mean() / close
        features["vol_of_vol"] = features["vol_20"].rolling(20).std()
        features["price_zscore"] = (close - close.rolling(50).mean()) / (close.rolling(50).std() + eps)
        vwap = (close * volume).rolling(20).sum() / (volume.rolling(20).sum() + eps)
        features["dist_vwap"] = close / vwap - 1.0
        obv = (np.sign(close.diff()) * volume).cumsum()
        features["obv_slope"] = obv.diff(10) / (obv.rolling(10).std() + eps)
        features["ret_skew_20"] = features["ret_1"].rolling(20).skew()
        features["ret_kurt_20"] = features["ret_1"].rolling(20).kurt()
        plus_dm = (high - high.shift(1)).clip(lower=0)
        minus_dm = (low.shift(1) - low).clip(lower=0)
        tr_smooth = tr.rolling(14).mean()
        plus_di = 100 * plus_dm.rolling(14).mean() / (tr_smooth + eps)
        minus_di = 100 * minus_dm.rolling(14).mean() / (tr_smooth + eps)
        features["trend_strength"] = (plus_di - minus_di).abs() / (plus_di + minus_di + eps)
    
    X = pd.DataFrame(features, index=df.index)
    fwd_ret = features["ret_1"].shift(-1)
    y = (fwd_ret > 0).astype(float) 
    
    price_data = pd.DataFrame({
        "close": close, "high": high, "low": low, "open": open_price, "vol_20_lagged": features["vol_20_lagged"] 
    }, index=df.index)
    
    X_clean = X.replace([np.inf, -np.inf], np.nan)
    valid_features = X_clean.notna().all(axis=1)
    valid_target = fwd_ret.notna()
    valid_mask = valid_features & valid_target
    
    X_out = X_clean.loc[valid_mask].copy()
    y_out = y.loc[valid_mask].astype(int)
    fwd_ret_out = fwd_ret.loc[valid_mask].copy()
    price_data_out = price_data.loc[valid_mask].copy()
    
    return X_out, y_out, fwd_ret_out, price_data_out

def select_features_in_fold(X_train: pd.DataFrame, y_train: pd.Series, X_val: pd.DataFrame, X_test: pd.DataFrame, n_features: int = 12):
    mi_scores = mutual_info_classif(X_train.values, y_train.values, random_state=42)
    mi_df = pd.DataFrame({'feature': X_train.columns, 'mi_score': mi_scores}).sort_values('mi_score', ascending=False)
    top_features = mi_df.head(n_features)['feature'].tolist()
    return X_train[top_features], X_val[top_features], X_test[top_features], mi_df

# ===========================================================================
# STATISTICAL SIGNIFICANCE
# ===========================================================================

def bootstrap_confidence_intervals(log_returns: np.ndarray, n_bootstrap: int = 5000, ci: float = 0.95, seed: int = 42):
    rng = np.random.default_rng(seed)
    returns = np.asarray(log_returns)
    returns = returns[np.isfinite(returns)]
    n = len(returns)
    if n < 50: return {k: (np.nan, (np.nan, np.nan)) for k in ["sharpe", "cagr", "max_dd", "sortino"]}
    
    block_size = min(20, max(5, n // 20))
    sharpes, cagrs, max_dds, sortinos = [], [], [], []
    
    for _ in range(n_bootstrap):
        n_blocks = int(np.ceil(n / block_size))
        block_starts = rng.integers(0, max(1, n - block_size + 1), size=n_blocks)
        sample = np.concatenate([returns[start:min(start + block_size, n)] for start in block_starts])[:n]
        if len(sample) < 20 or sample.std() == 0: continue
        
        sharpes.append((sample.mean() / sample.std()) * np.sqrt(252))
        eq = np.exp(np.cumsum(sample))
        years = len(sample) / 252
        if years > 0 and eq[-1] > 0: cagrs.append(eq[-1] ** (1/years) - 1)
        dd = eq / np.maximum.accumulate(eq) - 1
        max_dds.append(dd.min())
        downside = sample[sample < 0]
        if len(downside) > 1 and downside.std() > 0:
            sortinos.append((sample.mean() * 252) / (downside.std() * np.sqrt(252)))
    
    alpha = (1 - ci) / 2
    def ci_stats(arr):
        arr = [x for x in arr if np.isfinite(x)]
        if len(arr) < 100: return np.nan, (np.nan, np.nan)
        return np.mean(arr), (np.percentile(arr, alpha*100), np.percentile(arr, (1-alpha)*100))
    
    return {"sharpe": ci_stats(sharpes), "cagr": ci_stats(cagrs), "max_dd": ci_stats(max_dds), "sortino": ci_stats(sortinos)}

def probabilistic_sharpe_ratio(observed_sharpe: float, benchmark_sharpe: float, n_returns: int, skewness: float, excess_kurtosis: float):
    if n_returns < 10 or not np.isfinite(observed_sharpe): return np.nan, np.nan, np.nan
    skew = np.clip(skewness, -10, 10) if np.isfinite(skewness) else 0
    ex_kurt = np.clip(excess_kurtosis, -10, 100) if np.isfinite(excess_kurtosis) else 0
    sr = observed_sharpe
    variance_sr = (1.0 / (n_returns - 1)) * (1.0 + 0.5 * sr**2 - skew * sr + (ex_kurt / 4.0) * sr**2)
    if variance_sr <= 0: return np.nan, np.nan, np.nan
    se = np.sqrt(variance_sr)
    z = (observed_sharpe - benchmark_sharpe) / se
    psr = scipy_stats.norm.cdf(z)
    return psr, z, se

def deflated_sharpe_ratio(observed_sharpe: float, n_returns: int, n_trials: int, skewness: float, excess_kurtosis: float):
    if n_trials < 1 or n_returns < 10: return np.nan, np.nan
    gamma_em = 0.5772156649
    e_max = 0 if n_trials == 1 else (1 - gamma_em) * scipy_stats.norm.ppf(1 - 1/n_trials) + gamma_em * scipy_stats.norm.ppf(1 - 1/(n_trials * np.e))
    psr, _, _ = probabilistic_sharpe_ratio(observed_sharpe, e_max, n_returns, skewness, excess_kurtosis)
    return psr, e_max

def compute_statistical_significance(log_returns: np.ndarray, n_trials_tested: int = 1, benchmark_sharpe: float = 0.0):
    returns = np.asarray(log_returns)
    returns = returns[np.isfinite(returns)]
    n = len(returns)
    if n < 50: return {"error": "Insufficient data for statistical tests"}
    
    sharpe = (returns.mean() / returns.std()) * np.sqrt(252) if returns.std() > 0 else np.nan
    skew = scipy_stats.skew(returns)
    excess_kurt = scipy_stats.kurtosis(returns, fisher=True)
    
    bootstrap_ci = bootstrap_confidence_intervals(returns)
    psr, psr_z, psr_se = probabilistic_sharpe_ratio(sharpe, benchmark_sharpe, n, skew, excess_kurt)
    dsr, expected_max_sharpe = deflated_sharpe_ratio(sharpe, n, n_trials_tested, skew, excess_kurt)
    t_stat, t_pvalue = scipy_stats.ttest_1samp(returns, 0)
    
    return {
        "n_observations": n, "sharpe_point_estimate": sharpe, "sharpe_95_ci": bootstrap_ci["sharpe"][1],
        "cagr_95_ci": bootstrap_ci["cagr"][1], "max_dd_95_ci": bootstrap_ci["max_dd"][1], "sortino_95_ci": bootstrap_ci["sortino"][1],
        "skewness": skew, "excess_kurtosis": excess_kurt, "psr_vs_zero": psr, "psr_z_score": psr_z, "psr_std_error": psr_se,
        "deflated_sharpe_ratio": dsr, "expected_max_sharpe_null": expected_max_sharpe, "n_trials_tested": n_trials_tested,
        "t_statistic": t_stat, "t_pvalue": t_pvalue, "significant_at_5pct": t_pvalue < 0.05 if np.isfinite(t_pvalue) else False,
    }

# ===========================================================================
# HMM REGIME DETECTION
# ===========================================================================

class StudentTHMM:
    def __init__(self, n_components: int = 3, n_iter: int = 100, tol: float = 1e-3, df: float = 5.0, random_state: int = 42):
        self.n_components = n_components
        self.n_iter = n_iter
        self.tol = tol
        self.df = df
        self.random_state = random_state
        self.startprob_, self.transmat_, self.means_, self.covars_, self.cholesky_factors_ = None, None, None, None, None
        self.monitor_ = type('obj', (object,), {'converged': True})()
    
    def _compute_cholesky(self, cov: np.ndarray) -> np.ndarray:
        n = cov.shape[0]
        reg = 1e-6
        for _ in range(10):
            try:
                return cholesky(cov + np.eye(n) * reg, lower=True)
            except np.linalg.LinAlgError:
                reg *= 10
        return np.diag(np.sqrt(np.diag(cov) + 1e-4))
    
    def _mahalanobis_cholesky(self, X: np.ndarray, mean: np.ndarray, L: np.ndarray) -> np.ndarray:
        diff = X - mean  
        try:
            z = solve_triangular(L, diff.T, lower=True)  
            return np.sum(z**2, axis=0)  
        except:
            return np.sum(diff**2, axis=1)
    
    def _log_det_cholesky(self, L: np.ndarray) -> float:
        return 2.0 * np.sum(np.log(np.diag(L) + 1e-10))
    
    def fit(self, X: np.ndarray):
        np.random.seed(self.random_state)
        n_samples, n_features = X.shape
        k = self.n_components
        
        from sklearn.cluster import KMeans
        kmeans = KMeans(n_clusters=k, random_state=self.random_state, n_init=10)
        labels = kmeans.fit_predict(X)
        
        self.startprob_ = np.bincount(labels[:min(100, n_samples)], minlength=k).astype(float)
        self.startprob_ = (self.startprob_ + 1) / (self.startprob_.sum() + k)
        
        self.transmat_ = np.ones((k, k)) / k
        for i in range(len(labels) - 1): self.transmat_[labels[i], labels[i+1]] += 1
        self.transmat_ = self.transmat_ / self.transmat_.sum(axis=1, keepdims=True)
        
        self.means_ = kmeans.cluster_centers_.copy()
        self.covars_ = np.array([np.cov(X[labels == i].T) + np.eye(n_features) * 1e-4 if (labels == i).sum() > n_features else np.eye(n_features) * 0.1 for i in range(k)])
        self.cholesky_factors_ = [self._compute_cholesky(self.covars_[i]) for i in range(k)]
        
        prev_ll = -np.inf
        for _ in range(self.n_iter):
            log_resp = self._compute_log_likelihood(X)
            resp = np.exp(log_resp - log_resp.max(axis=1, keepdims=True))
            resp = resp / (resp.sum(axis=1, keepdims=True) + 1e-10)
            
            for ki in range(k):
                weights = resp[:, ki]
                if weights.sum() < 1e-6: continue
                maha = self._mahalanobis_cholesky(X, self.means_[ki], self.cholesky_factors_[ki])
                u = (self.df + n_features) / (self.df + maha + 1e-8)
                effective_weights = weights * u
                
                if effective_weights.sum() > 1e-6:
                    self.means_[ki] = np.average(X, weights=effective_weights, axis=0)
                    diff = X - self.means_[ki]
                    self.covars_[ki] = np.average(diff[:, :, np.newaxis] * diff[:, np.newaxis, :], weights=effective_weights, axis=0) + np.eye(n_features) * 1e-4
                    self.cholesky_factors_[ki] = self._compute_cholesky(self.covars_[ki])
            
            ll = log_resp.max(axis=1).sum()
            if abs(ll - prev_ll) < self.tol: break
            prev_ll = ll
        return self
    
    def _compute_log_likelihood(self, X: np.ndarray) -> np.ndarray:
        n_samples, n_features = X.shape
        log_prob = np.zeros((n_samples, self.n_components))
        for k in range(self.n_components):
            maha = self._mahalanobis_cholesky(X, self.means_[k], self.cholesky_factors_[k])
            log_det = self._log_det_cholesky(self.cholesky_factors_[k])
            log_prob[:, k] = (scipy_stats.gammaln((self.df + n_features) / 2) - scipy_stats.gammaln(self.df / 2) - (n_features / 2) * np.log(self.df * np.pi) - 0.5 * log_det - ((self.df + n_features) / 2) * np.log(1 + maha / self.df))
        return log_prob
    
    def predict_proba(self, X: np.ndarray) -> np.ndarray:
        log_prob = self._compute_log_likelihood(X)
        prob = np.exp(log_prob - log_prob.max(axis=1, keepdims=True))
        return prob / (prob.sum(axis=1, keepdims=True) + 1e-10)
    
    def predict(self, X: np.ndarray) -> np.ndarray:
        return self.predict_proba(X).argmax(axis=1)

def fit_hmm(X_train: np.ndarray, n_states: int = 3, random_state: int = 42, use_student_t: bool = False, max_retries: int = 3):
    if use_student_t:
        model = StudentTHMM(n_components=n_states, random_state=random_state)
        model.fit(X_train)
        return model
    else:
        for attempt in range(max_retries):
            model = GaussianHMM(n_components=n_states, covariance_type="full", n_iter=300, tol=1e-3, random_state=random_state + attempt, min_covar=1e-4)
            model.fit(X_train)
            if model.monitor_.converged: return model
        return model

def regime_churn(state_sequence: np.ndarray, window: int = 20, max_flips: int = 6) -> np.ndarray:
    states = np.asarray(state_sequence)
    flips = np.concatenate([[0], (np.diff(states) != 0).astype(int)])
    rolling_flips = pd.Series(flips).rolling(window, min_periods=1).sum().values
    return rolling_flips <= max_flips

def weighted_state_scores(state_probs: np.ndarray, forward_returns: np.ndarray) -> Dict[int, float]:
    scores = {}
    r = np.asarray(forward_returns)
    for s in range(state_probs.shape[1]):
        w = state_probs[:, s]
        mask = np.isfinite(r) & np.isfinite(w)
        if mask.sum() == 0 or w[mask].sum() == 0: scores[s] = np.nan
        else: scores[s] = np.average(r[mask], weights=w[mask])
    return scores

# ===========================================================================
# NEURAL NETWORK
# ===========================================================================

class SequentialDataset(Dataset):
    def __init__(self, X: np.ndarray, y: np.ndarray, seq_len: int = 20):
        self.X = torch.tensor(X, dtype=torch.float32)
        self.y = torch.tensor(y.reshape(-1, 1), dtype=torch.float32)
        self.seq_len = seq_len
    def __len__(self): return max(0, len(self.X) - self.seq_len + 1)
    def __getitem__(self, idx): return self.X[idx:idx + self.seq_len], self.y[idx + self.seq_len - 1]

class HybridMLP(nn.Module):
    def __init__(self, input_dim: int, hidden_dim: int = 64, dropout: float = 0.25, **kwargs):
        super().__init__()
        self.net = nn.Sequential(
            nn.Linear(input_dim, hidden_dim), nn.BatchNorm1d(hidden_dim), nn.ReLU(), nn.Dropout(dropout),
            nn.Linear(hidden_dim, hidden_dim // 2), nn.BatchNorm1d(hidden_dim // 2), nn.ReLU(), nn.Dropout(dropout),
            nn.Linear(hidden_dim // 2, 1),
        )
    def forward(self, x):
        if x.dim() == 3: x = x[:, -1, :]  
        return self.net(x)

class HybridResNet(nn.Module):
    def __init__(self, input_dim: int, hidden_dim: int = 64, dropout: float = 0.25, num_layers: int = 3, **kwargs):
        super().__init__()
        self.input_proj = nn.Linear(input_dim, hidden_dim)
        self.input_bn = nn.BatchNorm1d(hidden_dim)
        self.blocks = nn.ModuleList([nn.Sequential(nn.Linear(hidden_dim, hidden_dim), nn.BatchNorm1d(hidden_dim), nn.ReLU(), nn.Dropout(dropout), nn.Linear(hidden_dim, hidden_dim), nn.BatchNorm1d(hidden_dim)) for _ in range(num_layers)])
        self.fc = nn.Linear(hidden_dim, 1)
    def forward(self, x):
        if x.dim() == 3: x = x[:, -1, :]
        x = F.relu(self.input_bn(self.input_proj(x)))
        for block in self.blocks: x = F.relu(x + block(x))
        return self.fc(x)

class HybridLSTM(nn.Module):
    def __init__(self, input_dim: int, hidden_dim: int = 64, dropout: float = 0.25, num_layers: int = 2, seq_len: int = 20, **kwargs):
        super().__init__()
        self.seq_len = seq_len
        self.lstm = nn.LSTM(input_size=input_dim, hidden_size=hidden_dim, num_layers=num_layers, batch_first=True, dropout=dropout if num_layers > 1 else 0)
        self.attention = nn.Sequential(nn.Linear(hidden_dim, hidden_dim // 2), nn.Tanh(), nn.Linear(hidden_dim // 2, 1))
        self.fc = nn.Sequential(nn.Linear(hidden_dim, hidden_dim // 2), nn.ReLU(), nn.Dropout(dropout), nn.Linear(hidden_dim // 2, 1))
    def forward(self, x):
        if x.dim() == 2: x = x.unsqueeze(1).expand(-1, self.seq_len, -1)
        lstm_out, _ = self.lstm(x)  
        attn_weights = F.softmax(self.attention(lstm_out), dim=1)
        context = torch.sum(attn_weights * lstm_out, dim=1)
        return self.fc(context)

class HybridTransformer(nn.Module):
    def __init__(self, input_dim: int, hidden_dim: int = 64, dropout: float = 0.25, num_layers: int = 2, seq_len: int = 20, nhead: int = 4, **kwargs):
        super().__init__()
        self.seq_len = seq_len
        if hidden_dim % nhead != 0: hidden_dim = (hidden_dim // nhead) * nhead
        self.input_proj = nn.Linear(input_dim, hidden_dim)
        self.pos_encoding = nn.Parameter(torch.randn(1, seq_len, hidden_dim) * 0.1)
        encoder_layer = nn.TransformerEncoderLayer(d_model=hidden_dim, nhead=nhead, dim_feedforward=hidden_dim * 4, dropout=dropout, batch_first=True, activation='gelu')
        self.transformer = nn.TransformerEncoder(encoder_layer, num_layers=num_layers)
        self.fc = nn.Sequential(nn.Linear(hidden_dim, hidden_dim // 2), nn.GELU(), nn.Dropout(dropout), nn.Linear(hidden_dim // 2, 1))
    def forward(self, x):
        if x.dim() == 2: x = x.unsqueeze(1).expand(-1, self.seq_len, -1)
        x = self.input_proj(x)
        x = x + self.pos_encoding[:, :x.size(1), :]
        x = self.transformer(x)
        return self.fc(x[:, -1, :])

def get_model_class(architecture: str):
    return {"mlp": HybridMLP, "resnet": HybridResNet, "lstm": HybridLSTM, "transformer": HybridTransformer}.get(architecture.lower(), HybridMLP)

def is_sequential_model(architecture: str) -> bool:
    return architecture.lower() in ["lstm", "transformer"]

def _val_sharpe(model: nn.Module, X_val: np.ndarray, y_val_ret: np.ndarray, device: torch.device, prob_long: float, prob_short: float, is_sequential: bool = False, seq_len: int = 20) -> float:
    model.eval()
    with torch.no_grad():
        if is_sequential:
            if len(X_val) < seq_len: return -np.inf
            n_seq = len(X_val) - seq_len + 1
            X_seq = np.array([X_val[i:i+seq_len] for i in range(n_seq)])
            X_t = torch.tensor(X_seq, dtype=torch.float32).to(device)
            logits = model(X_t).cpu().numpy().ravel()
            y_ret_aligned = y_val_ret[seq_len-1:]
        else:
            X_t = torch.tensor(X_val, dtype=torch.float32).to(device)
            logits = model(X_t).cpu().numpy().ravel()
            y_ret_aligned = y_val_ret
    prob = 1.0 / (1.0 + np.exp(-logits))
    sig = np.where(prob >= prob_long, 1, np.where(prob <= prob_short, -1, 0))
    min_len = min(len(sig), len(y_ret_aligned))
    r = sig[:min_len] * np.asarray(y_ret_aligned)[:min_len]
    if len(r) == 0 or r.std() == 0: return -np.inf
    return (r.mean() / r.std()) * np.sqrt(252)

def train_nn(X_train: np.ndarray, y_train: np.ndarray, X_val: np.ndarray, y_val: np.ndarray, y_val_ret: np.ndarray, prob_long: float, prob_short: float, architecture: str = "mlp", epochs: int = 120, batch_size: int = 64, lr: float = 1e-3, hidden_dim: int = 64, dropout: float = 0.25, weight_decay: float = 1e-4, patience: int = 15, clip_grad: float = 1.0, num_layers: int = 2, seq_len: int = 20, use_mixed_precision: bool = False):
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    is_sequential = is_sequential_model(architecture)
    if is_sequential and len(X_train) < seq_len + 10: architecture, is_sequential = "mlp", False
    
    if is_sequential:
        dataset = SequentialDataset(X_train, y_train, seq_len=seq_len)
        if len(dataset) == 0: architecture, is_sequential = "mlp", False
    if not is_sequential:
        dataset = TensorDataset(torch.tensor(X_train, dtype=torch.float32), torch.tensor(y_train.reshape(-1, 1), dtype=torch.float32))
    
    effective_batch = min(batch_size, len(dataset))
    if effective_batch == 0: raise ValueError("No training data available")
    loader = DataLoader(dataset, batch_size=effective_batch, shuffle=False, drop_last=len(dataset) > effective_batch)
    
    model = get_model_class(architecture)(input_dim=X_train.shape[1], hidden_dim=hidden_dim, dropout=dropout, num_layers=num_layers, seq_len=seq_len).to(device)
    optimizer = torch.optim.Adam(model.parameters(), lr=lr, weight_decay=weight_decay)
    criterion = nn.BCEWithLogitsLoss()
    scaler = torch.cuda.amp.GradScaler() if use_mixed_precision and device.type == "cuda" else None
    
    best_sharpe = -np.inf
    best_state = None
    wait = 0
    
    for _ in range(epochs):
        model.train()
        for xb, yb in loader:
            xb, yb = xb.to(device), yb.to(device)
            optimizer.zero_grad()
            if scaler is not None:
                with torch.cuda.amp.autocast(): loss = criterion(model(xb), yb)
                scaler.scale(loss).backward()
                scaler.unscale_(optimizer)
                nn.utils.clip_grad_norm_(model.parameters(), clip_grad)
                scaler.step(optimizer)
                scaler.update()
            else:
                loss = criterion(model(xb), yb)
                loss.backward()
                nn.utils.clip_grad_norm_(model.parameters(), clip_grad)
                optimizer.step()
        
        vs = _val_sharpe(model, X_val, y_val_ret, device, prob_long, prob_short, is_sequential, seq_len)
        if vs > best_sharpe:
            best_sharpe, best_state, wait = vs, copy.deepcopy(model.state_dict()), 0
        else:
            wait += 1
            if wait >= patience: break
            
    if best_state is not None: model.load_state_dict(best_state)
    return model

def predict_proba(model: nn.Module, X: np.ndarray, is_sequential: bool = False, seq_len: int = 20) -> np.ndarray:
    device = next(model.parameters()).device
    model.eval()
    with torch.no_grad():
        if is_sequential and len(X) >= seq_len:
            n_seq = len(X) - seq_len + 1
            X_seq = np.array([X[i:i+seq_len] for i in range(n_seq)])
            X_t = torch.tensor(X_seq, dtype=torch.float32).to(device)
            logits = model(X_t).cpu().numpy().ravel()
            full_probs = np.full(len(X), 0.5)
            full_probs[seq_len-1:] = 1.0 / (1.0 + np.exp(-logits))
            return full_probs
        else:
            X_t = torch.tensor(X, dtype=torch.float32).to(device)
            logits = model(X_t).cpu().numpy().ravel()
            return 1.0 / (1.0 + np.exp(-logits))

def tune_hyperparams(X_train: np.ndarray, y_train: np.ndarray, fwd_train: np.ndarray, prob_long: float, prob_short: float, regime_gate: float, churn_window: int, max_churn_flips: int, n_trials: int = 15, inner_splits: int = 3, seed: int = 42, use_student_t_hmm: bool = False, seq_len: int = 20, use_mixed_precision: bool = False, progress_callback=None):
    inner_tscv = TimeSeriesSplit(n_splits=inner_splits)
    split_data = []
    for itr_idx, ival_idx in inner_tscv.split(X_train):
        if len(itr_idx) < 300 or len(ival_idx) < 20:
            split_data.append(None)
            continue
        scaler_i = StandardScaler()
        Xtr = scaler_i.fit_transform(X_train[itr_idx])
        Xvl = scaler_i.transform(X_train[ival_idx])
        split_data.append((itr_idx, ival_idx, Xtr, Xvl, scaler_i))
    
    hmm_cache = {n: [] for n in [2, 3, 4]}
    for n_states in [2, 3, 4]:
        for s_data in split_data:
            if s_data is None:
                hmm_cache[n_states].append(None)
                continue
            itr_idx, ival_idx, Xtr, Xvl, scaler_i = s_data
            try:
                hmm_i = fit_hmm(Xtr, n_states=n_states, random_state=7, use_student_t=use_student_t_hmm)
                Ptr, Pvl = hmm_i.predict_proba(Xtr), hmm_i.predict_proba(Xvl)
                stable_m = regime_churn(hmm_i.predict(Xvl), window=churn_window, max_flips=max_churn_flips)
                ss_i = {k: v for k, v in weighted_state_scores(Ptr, fwd_train[itr_idx]).items() if pd.notna(v)}
                if not ss_i:
                    hmm_cache[n_states].append(None)
                    continue
                b_state, w_state = max(ss_i, key=ss_i.get), min(ss_i, key=ss_i.get)
                if b_state == w_state: w_state = -1 
                hmm_cache[n_states].append((Ptr, Pvl, b_state, w_state, stable_m, scaler_i))
            except Exception:
                hmm_cache[n_states].append(None)

    def objective(trial):
        n_states = trial.suggest_categorical("n_states", [2, 3, 4])
        architecture = trial.suggest_categorical("architecture", ["mlp", "resnet"])
        hidden_dim = trial.suggest_categorical("hidden_dim", [32, 64])
        dropout = trial.suggest_categorical("dropout", [0.20, 0.35])
        lr = trial.suggest_categorical("lr", [5e-4, 1e-3])
        sharpes = []
        n_states_cache = hmm_cache[n_states]
        
        for split_idx, s_data in enumerate(split_data):
            if s_data is None or n_states_cache[split_idx] is None: continue
            itr_idx, ival_idx, Xtr, Xvl, scaler_i = s_data
            Ptr, Pvl, best_state, worst_state, stable_mask, _ = n_states_cache[split_idx]
            
            scaler_h = StandardScaler()
            Xtr_h = scaler_h.fit_transform(np.hstack([Xtr, Ptr]))
            Xvl_h = scaler_h.transform(np.hstack([Xvl, Pvl]))
            
            try:
                is_seq = is_sequential_model(architecture)
                m_i = train_nn(Xtr_h, y_train[itr_idx], Xvl_h, y_train[ival_idx], fwd_train[ival_idx], prob_long=prob_long, prob_short=prob_short, architecture=architecture, hidden_dim=hidden_dim, dropout=dropout, lr=lr, epochs=60, patience=8, seq_len=seq_len, use_mixed_precision=use_mixed_precision)
                prob_i = predict_proba(m_i, Xvl_h, is_seq, seq_len)
                valid_mask = ~np.isnan(prob_i)
                if valid_mask.sum() == 0: continue
                
                prob_best = Pvl[:, best_state] if best_state >= 0 else np.zeros(len(Pvl))
                prob_worst = Pvl[:, worst_state] if worst_state >= 0 else np.zeros(len(Pvl))
                raw_signal = np.where(valid_mask & (prob_i >= prob_long) & (prob_best >= regime_gate), 1, np.where(valid_mask & (prob_i <= prob_short) & (prob_worst >= regime_gate), -1, 0))
                sig_i = np.where(stable_mask, raw_signal, 0)
                valid_r = (sig_i * fwd_train[ival_idx])[valid_mask]
                if len(valid_r) > 0 and valid_r.std() > 0: sharpes.append(valid_r.mean() / valid_r.std() * np.sqrt(252))
            except Exception: continue
        return np.mean(sharpes) if sharpes else -np.inf

    sampler = optuna.samplers.TPESampler(seed=seed)
    study = optuna.create_study(direction="maximize", sampler=sampler)
    
    callbacks = []
    if progress_callback:
        callbacks.append(lambda s, t: progress_callback(t.number))
        
    study.optimize(objective, n_trials=n_trials, callbacks=callbacks)
    return study.best_params, study.best_value

# ===========================================================================
# RISK MANAGEMENT & TUNING & BENCHMARKS
# ===========================================================================

class RiskManager:
    def __init__(self, stop_loss_pct=0.02, take_profit_pct=0.05, trailing_stop_pct=0.015, max_drawdown_halt=0.10, use_conservative_gaps=True, regime_stop_multipliers=None):
        self.stop_loss_pct, self.take_profit_pct, self.trailing_stop_pct = stop_loss_pct, take_profit_pct, trailing_stop_pct
        self.max_drawdown_halt, self.use_conservative_gaps = max_drawdown_halt, use_conservative_gaps
        self.regime_stop_multipliers = regime_stop_multipliers or {}
        self.max_position = 1.5
        self.reset()
        
    def reset(self):
        self.entry_price, self.entry_signal, self.high_water_mark = None, 0, None
        self.current_equity, self.peak_equity, self.is_halted = 1.0, 1.0, False
        
    def _close_position(self): self.entry_price, self.entry_signal, self.high_water_mark = None, 0, None
    
    def apply_risk_rules(self, raw_signal, close_price, high_price, low_price, open_price, current_regime, daily_pnl=0.0):
        self.current_equity *= np.exp(daily_pnl)
        self.peak_equity = max(self.peak_equity, self.current_equity)
        if (self.peak_equity - self.current_equity) / self.peak_equity > self.max_drawdown_halt: self.is_halted = True; return 0.0, "max_drawdown_halt"
        if self.is_halted: return 0.0, "halted"
        
        stop_mult = self.regime_stop_multipliers.get(current_regime, 1.0)
        eff_stop = self.stop_loss_pct * stop_mult
        eff_trail = self.trailing_stop_pct * stop_mult
        
        if self.entry_price is not None and self.entry_signal != 0:
            if self.use_conservative_gaps:
                if self.entry_signal > 0: worst_price, best_price, pnl_pct_worst, pnl_pct_best = low_price, high_price, (low_price - self.entry_price) / self.entry_price, (high_price - self.entry_price) / self.entry_price
                else: worst_price, best_price, pnl_pct_worst, pnl_pct_best = high_price, low_price, (self.entry_price - high_price) / self.entry_price, (self.entry_price - low_price) / self.entry_price
            else:
                pnl_pct_worst = pnl_pct_best = (close_price - self.entry_price) / self.entry_price if self.entry_signal > 0 else (self.entry_price - close_price) / self.entry_price
            
            if self.high_water_mark is None: self.high_water_mark = self.entry_price
            else: self.high_water_mark = max(self.high_water_mark, best_price) if self.entry_signal > 0 else min(self.high_water_mark, best_price)
            
            if pnl_pct_worst < -eff_stop: self._close_position(); return 0.0, "stop_loss"
            if pnl_pct_best > self.take_profit_pct: self._close_position(); return 0.0, "take_profit"
            trailing_pnl = (worst_price - self.high_water_mark) / self.high_water_mark if self.entry_signal > 0 else (self.high_water_mark - worst_price) / self.high_water_mark
            if trailing_pnl < -eff_trail: self._close_position(); return 0.0, "trailing_stop"
            
        if raw_signal != 0 and self.entry_price is None:
            self.entry_price = open_price if self.use_conservative_gaps else close_price
            self.entry_signal = np.sign(raw_signal)
            self.high_water_mark = self.entry_price
        elif raw_signal == 0 and self.entry_price is not None: self._close_position()
        elif raw_signal != 0 and self.entry_price is not None and np.sign(raw_signal) != self.entry_signal:
            self._close_position()
            self.entry_price = open_price if self.use_conservative_gaps else close_price
            self.entry_signal, self.high_water_mark = np.sign(raw_signal), self.entry_price
            
        return np.clip(raw_signal, -self.max_position, self.max_position), None

def apply_risk_management(signals: np.ndarray, price_data: pd.DataFrame, regimes: np.ndarray, returns: np.ndarray, risk_config: RiskConfig):
    if not risk_config.enabled: return signals, [None] * len(signals)
    rm = RiskManager(stop_loss_pct=risk_config.stop_loss_pct, take_profit_pct=risk_config.take_profit_pct, trailing_stop_pct=risk_config.trailing_stop_pct, max_drawdown_halt=risk_config.max_drawdown_halt, use_conservative_gaps=risk_config.use_conservative_gaps, regime_stop_multipliers=risk_config.regime_stop_multipliers)
    adjusted_signals, exit_reasons = np.zeros_like(signals), []
    for i in range(len(signals)):
        daily_pnl = adjusted_signals[i-1] * returns[i-1] if i > 0 else 0
        adj_sig, reason = rm.apply_risk_rules(signals[i], price_data["close"].iloc[i], price_data["high"].iloc[i], price_data["low"].iloc[i], price_data["open"].iloc[i], regimes[i], daily_pnl)
        adjusted_signals[i] = adj_sig
        exit_reasons.append(reason)
    return adjusted_signals, exit_reasons

def compute_transaction_costs(signals: np.ndarray, vol_daily: np.ndarray, cost_bps: float = 2.0, spread_bps: float = 1.0, impact_factor: float = 0.1):
    trade_change = np.abs(np.diff(np.concatenate([[0], signals])))
    return trade_change * ((cost_bps / 10000.0) + (spread_bps / 10000.0) + (impact_factor * vol_daily))

def logistic_benchmark(X_train_scaled, y_train, X_test_scaled, fwd_test, prob_long, prob_short, vol_daily_lagged, vol_target, cost_bps, spread_bps, impact_factor):
    prob = LogisticRegression(max_iter=500, C=0.1, random_state=42).fit(X_train_scaled, y_train).predict_proba(X_test_scaled)[:, 1]
    signal = np.where(prob >= prob_long, 1, np.where(prob <= prob_short, -1, 0))
    if vol_target > 0 and vol_daily_lagged is not None: signal = signal * np.clip(vol_target / (vol_daily_lagged * np.sqrt(252) + 1e-8), 0.0, 1.5)
    return signal * fwd_test - compute_transaction_costs(signal, vol_daily_lagged if vol_daily_lagged is not None else np.zeros(len(signal)), cost_bps, spread_bps, impact_factor)

def monte_carlo_benchmark(fwd_test, vol_daily, cost_bps, spread_bps, impact_factor, n_sims=500, seed=0, avg_holding_period=5):
    rng = np.random.default_rng(seed)
    all_ret = []
    n = len(fwd_test)
    for _ in range(n_sims):
        sig, i = np.zeros(n), 0
        while i < n:
            direction, hold_days = rng.choice([-1, 0, 1]), rng.geometric(1/avg_holding_period)
            sig[i:min(i+hold_days, n)] = direction
            i += hold_days
        all_ret.append(sig * fwd_test - compute_transaction_costs(sig, vol_daily, cost_bps, spread_bps, impact_factor))
    arr = np.array(all_ret)
    return arr.mean(axis=0), arr.std(axis=0), arr[np.argsort(np.sum(arr, axis=1))[int(0.95 * n_sims)]]

def performance_stats(log_returns: np.ndarray, periods_per_year: int = 252, signals: np.ndarray = None):
    r = pd.Series(log_returns).dropna()
    if len(r) == 0: return {"cagr": np.nan, "vol": np.nan, "sharpe": np.nan, "sortino": np.nan, "max_dd": np.nan, "hit_rate": np.nan, "calmar": np.nan}
    eq = np.exp(r.cumsum())
    years = len(r) / periods_per_year
    cagr = eq.iloc[-1] ** (1 / years) - 1 if years > 0 else np.nan
    vol = r.std() * np.sqrt(periods_per_year)
    sharpe = (r.mean() * periods_per_year) / vol if vol > 0 else np.nan
    downside = r[r < 0]
    downside_vol = downside.std() * np.sqrt(periods_per_year) if len(downside) > 0 else np.nan
    sortino = (r.mean() * periods_per_year) / downside_vol if downside_vol > 0 else np.nan
    max_dd = (eq / eq.cummax() - 1).min()
    return {"cagr": cagr, "vol": vol, "sharpe": sharpe, "sortino": sortino, "max_dd": max_dd, "hit_rate": (r[signals != 0] > 0).mean() if signals is not None and len(r[signals != 0]) > 0 else (r > 0).mean(), "calmar": cagr / abs(max_dd) if max_dd != 0 else np.nan}

# ===========================================================================
# BACKTEST LOOP
# ===========================================================================

def run_hybrid_backtest(config: Config, progress_bar, status_text) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, Dict]:
    df = load_data(config.data.ticker, config.data.period, config.data.interval)
    X_df, y, fwd_ret, price_data = build_features(df, use_advanced=config.features.use_advanced_features)
    
    max_train = None if config.backtest.anchored else int(len(X_df) / (config.backtest.n_splits + 1)) * 2
    splitter = TimeSeriesSplit(n_splits=config.backtest.n_splits, max_train_size=max_train)
    fold_outputs = []
    hp_log = []
    
    splits = list(splitter.split(X_df))
    for fold, (train_idx, test_idx) in enumerate(splits, start=1):
        status_text.text(f"Processing Fold {fold}/{config.backtest.n_splits}...")
        progress_bar.progress((fold - 1) / config.backtest.n_splits)
        
        X_train_full, y_train_full, r_train_full = X_df.iloc[train_idx], y.iloc[train_idx], fwd_ret.iloc[train_idx]
        X_test_df, y_test, r_test, price_test = X_df.iloc[test_idx], y.iloc[test_idx], fwd_ret.iloc[test_idx], price_data.iloc[test_idx]
        if len(train_idx) < 300: continue
            
        split_point = int(len(X_train_full) * 0.85)
        X_train_df, y_train, r_train = X_train_full.iloc[:split_point], y_train_full.iloc[:split_point], r_train_full.iloc[:split_point]
        X_val_df, y_val, r_val = X_train_full.iloc[split_point:], y_train_full.iloc[split_point:], r_train_full.iloc[split_point:]
        
        if config.features.feature_selection:
            X_train_df, X_val_df, X_test_df, _ = select_features_in_fold(X_train_df, y_train, X_val_df, X_test_df, n_features=config.features.n_selected_features)
            
        if config.tuning.enabled and len(X_train_df) >= 300:
            status_text.text(f"Fold {fold}: Tuning Hyperparameters with Optuna...")
            
            def opt_cb(t_num): status_text.text(f"Fold {fold}: Tuning Optuna Trial {t_num}/{config.tuning.n_trials}")
            
            best_hp, best_hp_score = tune_hyperparams(X_train_df.values, y_train.values, r_train.values, config.trading.prob_long, config.trading.prob_short, config.trading.regime_gate, config.trading.churn_window, config.trading.max_churn, n_trials=config.tuning.n_trials, inner_splits=config.tuning.inner_splits, seed=config.backtest.seed, use_student_t_hmm=config.model.use_student_t_hmm, seq_len=config.model.seq_len, use_mixed_precision=config.model.use_mixed_precision, progress_callback=opt_cb)
            n_states_fold, architecture_fold, hidden_dim_f, dropout_f, lr_f = best_hp["n_states"], best_hp.get("architecture", config.model.architecture), best_hp["hidden_dim"], best_hp["dropout"], best_hp["lr"]
        else:
            n_states_fold, architecture_fold, hidden_dim_f, dropout_f, lr_f = config.model.n_states, config.model.architecture, config.model.hidden_dim, config.model.dropout, config.model.lr
            best_hp_score = np.nan
            
        hp_log.append({"fold": fold, "n_states": n_states_fold, "architecture": architecture_fold, "hidden_dim": hidden_dim_f, "dropout": dropout_f, "lr": lr_f, "inner_sharpe": best_hp_score})
            
        scaler = StandardScaler()
        X_train_scaled, X_val_scaled, X_test_scaled = scaler.fit_transform(X_train_df), scaler.transform(X_val_df), scaler.transform(X_test_df)
        
        status_text.text(f"Fold {fold}: Fitting HMM ({n_states_fold} states)...")
        hmm = fit_hmm(X_train_scaled, n_states=n_states_fold, random_state=config.backtest.seed + fold, use_student_t=config.model.use_student_t_hmm)
        train_sp, val_sp, test_sp = hmm.predict_proba(X_train_scaled), hmm.predict_proba(X_val_scaled), hmm.predict_proba(X_test_scaled)
        test_state_seq = hmm.predict(X_test_scaled)
        stable_mask = regime_churn(test_state_seq, window=config.trading.churn_window, max_flips=config.trading.max_churn)
        
        state_scores = {k: v for k, v in weighted_state_scores(train_sp, r_train.values).items() if pd.notna(v)}
        if not state_scores: continue
        best_state, worst_state = max(state_scores, key=state_scores.get), min(state_scores, key=state_scores.get)
        if best_state == worst_state: worst_state = -1 
        
        scaler_h = StandardScaler()
        X_train_h = scaler_h.fit_transform(np.hstack([X_train_scaled, train_sp]))
        X_val_h = scaler_h.transform(np.hstack([X_val_scaled, val_sp]))
        X_test_h = scaler_h.transform(np.hstack([X_test_scaled, test_sp]))
        
        status_text.text(f"Fold {fold}: Training PyTorch Neural Network ({architecture_fold})...")
        is_seq = is_sequential_model(architecture_fold)
        model = train_nn(X_train_h, y_train.values, X_val_h, y_val.values, r_val.values, prob_long=config.trading.prob_long, prob_short=config.trading.prob_short, architecture=architecture_fold, hidden_dim=hidden_dim_f, dropout=dropout_f, lr=lr_f, num_layers=config.model.num_layers, seq_len=config.model.seq_len, use_mixed_precision=config.model.use_mixed_precision)
        
        prob_up = predict_proba(model, X_test_h, is_seq, config.model.seq_len)
        valid_pred = ~np.isnan(prob_up)
        prob_up = np.where(valid_pred, prob_up, 0.5)
        prob_best = test_sp[:, best_state] if best_state >= 0 else np.zeros(len(test_sp))
        prob_worst = test_sp[:, worst_state] if worst_state >= 0 else np.zeros(len(test_sp))
        
        raw_signal = np.where(valid_pred & (prob_up >= config.trading.prob_long) & (prob_best >= config.trading.regime_gate), 1, np.where(valid_pred & (prob_up <= config.trading.prob_short) & (prob_worst >= config.trading.regime_gate), -1, 0))
        signal = np.where(stable_mask, raw_signal, 0)
        
        vol_daily_lagged = price_test["vol_20_lagged"].values
        sized_signal = signal * np.clip(config.trading.vol_target / (vol_daily_lagged * np.sqrt(252) + 1e-8), 0.0, 1.5) if config.trading.vol_target > 0 else signal.astype(float)
        
        if config.risk.enabled: sized_signal, exit_reasons = apply_risk_management(sized_signal, price_test, test_state_seq, r_test.values, config.risk)
        else: exit_reasons = [None] * len(sized_signal)
        
        costs = compute_transaction_costs(sized_signal, vol_daily_lagged, config.trading.cost_bps, config.trading.spread_bps, config.trading.impact_factor)
        strat_ret = sized_signal * r_test.values - costs
        
        lr_ret = logistic_benchmark(X_train_scaled, y_train.values, X_test_scaled, r_test.values, prob_long=config.trading.prob_long, prob_short=config.trading.prob_short, vol_daily_lagged=vol_daily_lagged, vol_target=config.trading.vol_target, cost_bps=config.trading.cost_bps, spread_bps=config.trading.spread_bps, impact_factor=config.trading.impact_factor)
        mc_mean, mc_std, mc_p95 = monte_carlo_benchmark(r_test.values, vol_daily_lagged, config.trading.cost_bps, config.trading.spread_bps, config.trading.impact_factor, seed=config.backtest.seed)
        
        fold_outputs.append(pd.DataFrame({"prob_up": prob_up, "prob_best_regime": prob_best, "prob_worst_regime": prob_worst, "stable": stable_mask.astype(int), "current_regime": test_state_seq, "signal": sized_signal, "exit_reason": exit_reasons, "strategy_log_ret": strat_ret, "buyhold_log_ret": r_test.values, "logistic_log_ret": lr_ret, "mc_mean_log_ret": mc_mean, "mc_p95_log_ret": mc_p95, "best_state": best_state, "worst_state": worst_state, "fold": fold}, index=X_test_df.index))

    progress_bar.progress(1.0)
    status_text.text("Backtest Complete! Compiling Results...")
    if not fold_outputs: raise RuntimeError("Backtest yielded no valid output folds.")
    
    results = pd.concat(fold_outputs).dropna(subset=["strategy_log_ret"])
    for col in ["strategy", "buyhold", "logistic", "mc_mean", "mc_p95"]:
        eq_col = col.replace("mean", "") + "_equity" if "mc_mean" in col else col + "_equity"
        results[eq_col] = np.exp(results[f"{col}_log_ret"].cumsum())
        if col in ["strategy", "buyhold", "logistic"]: results[f"dd_{col}"] = results[eq_col] / results[eq_col].cummax() - 1
            
    summary = pd.DataFrame([
        {"model": "Hybrid Markov + NN", **performance_stats(results["strategy_log_ret"], signals=results["signal"].values)},
        {"model": "Buy & Hold", **performance_stats(results["buyhold_log_ret"])},
        {"model": "Logistic Regression", **performance_stats(results["logistic_log_ret"])},
        {"model": "Monte Carlo Mean", **performance_stats(results["mc_mean_log_ret"])},
        {"model": "Monte Carlo P95", **performance_stats(results["mc_p95_log_ret"])}
    ])
    
    monthly = results[["strategy_log_ret", "buyhold_log_ret", "logistic_log_ret", "mc_mean_log_ret", "mc_p95_log_ret"]].copy()
    monthly.index = pd.to_datetime(monthly.index)
    monthly = monthly.resample("ME").sum().apply(lambda x: np.exp(x) - 1)
    monthly.columns = [c.replace("log_ret", "return") for c in monthly.columns]
    
    total_trials = config.backtest.n_splits if config.tuning.enabled else 1
    stat_tests = compute_statistical_significance(results["strategy_log_ret"].values, n_trials_tested=total_trials, benchmark_sharpe=0.0)
    
    return results, summary, monthly, pd.DataFrame(hp_log), stat_tests

def run_portfolio_backtest(tickers: List[str], config: Config, progress_bar, status_text, weights: Optional[np.ndarray] = None, correlation_threshold: float = 0.7) -> Dict[str, Any]:
    n_assets = len(tickers)
    if weights is None: weights = np.ones(n_assets) / n_assets
    
    all_results = {}
    for i, ticker in enumerate(tickers):
        progress_val = i / n_assets
        progress_bar.progress(progress_val)
        status_text.text(f"Running individual model for {ticker} ({i+1}/{n_assets})...")
        
        try:
            ticker_config = copy.deepcopy(config)
            ticker_config.data.ticker = ticker
            results, summary, monthly, hp_df, stats = run_hybrid_backtest(ticker_config, st.empty(), st.empty())
            all_results[ticker] = {'results': results, 'summary': summary, 'returns': results['strategy_log_ret'], 'stats': stats}
        except Exception as e:
            st.warning(f"Failed for {ticker}: {e}")
            continue
            
    progress_bar.progress(1.0)
    status_text.text("Portfolio Complete! Aggregating...")
            
    if len(all_results) < 2: raise ValueError("Need at least 2 successful backtests for portfolio calculation.")
    
    common_dates = None
    for ticker, data in all_results.items():
        dates = set(data['returns'].index)
        if common_dates is None: common_dates = dates
        else: common_dates = common_dates.intersection(dates)
            
    common_dates = sorted(common_dates)
    returns_matrix = pd.DataFrame({ticker: data['returns'].loc[common_dates] for ticker, data in all_results.items()})
    corr_matrix = returns_matrix.corr()
    
    adjusted_weights = weights[:len(returns_matrix.columns)].copy()
    for i, ticker_i in enumerate(returns_matrix.columns):
        for j, ticker_j in enumerate(returns_matrix.columns):
            if i != j and abs(corr_matrix.loc[ticker_i, ticker_j]) > correlation_threshold:
                adjusted_weights[i] *= 0.8
                
    adjusted_weights = adjusted_weights / adjusted_weights.sum()
    portfolio_returns = (returns_matrix * adjusted_weights).sum(axis=1)
    
    portfolio_stats = performance_stats(portfolio_returns)
    portfolio_stats['n_assets'] = len(all_results)
    portfolio_stats['avg_correlation'] = corr_matrix.values[np.triu_indices(len(corr_matrix), k=1)].mean()
    portfolio_stat_tests = compute_statistical_significance(portfolio_returns.values)
    
    return {'portfolio_returns': portfolio_returns, 'portfolio_stats': portfolio_stats, 'portfolio_stat_tests': portfolio_stat_tests, 'individual_results': all_results, 'correlation_matrix': corr_matrix, 'weights': dict(zip(returns_matrix.columns, adjusted_weights))}


# ===========================================================================
# VISUALIZATIONS & DASHBOARDS
# ===========================================================================

COLORS = {"hybrid": "#01696f", "buyhold": "#7a39bb", "logistic": "#da7101", "mc": "#bab9b4", "mcp95": "#8d99ae", "prob_up": "#006494", "regime": "#437a22", "signal": "#d19900", "dd_hybrid": "#a12c7b", "dd_bh": "#964219", "dd_lr": "#da7101"}

def build_single_asset_figure(results: pd.DataFrame, summary: pd.DataFrame, hp_df: pd.DataFrame, stat_tests: Dict[str, Any]):
    r = results.reset_index().copy()
    r.columns = [str(c) for c in r.columns]
    if 'index' in r.columns: r = r.rename(columns={'index': 'date'})
    if 'Date' in r.columns: r = r.rename(columns={'Date': 'date'})
    
    fig = make_subplots(
        rows=5, cols=2, shared_xaxes=False, vertical_spacing=0.07, horizontal_spacing=0.06,
        row_heights=[0.22, 0.18, 0.18, 0.22, 0.20], column_widths=[0.65, 0.35],
        subplot_titles=("Cumulative Equity Curves", "Performance Summary", "Strategy Drawdowns (Underwater)", "Monthly Return Distribution", "NN Probabilities & Trading Signals", "HMM Regime Stability", "Walk-Forward Hyperparameter Tuning Log", "Statistical Significance", "Rolling 63-Day Sharpe Ratio", "Annual Returns (%)"),
        specs=[[{"type": "xy"}, {"type": "domain"}], [{"type": "xy"}, {"type": "xy"}], [{"type": "xy"}, {"type": "xy"}], [{"type": "xy"}, {"type": "domain"}], [{"type": "xy"}, {"type": "xy"}]]
    )
    
    for col, name, color in [("strategy_equity", "Hybrid Markov + NN", COLORS["hybrid"]), ("buyhold_equity", "Buy & Hold", COLORS["buyhold"]), ("logistic_equity", "Logistic Regression", COLORS["logistic"]), ("mc_equity", "Monte Carlo Mean", COLORS["mc"]), ("mc_p95_equity", "Monte Carlo P95", COLORS["mcp95"])]:
        if col in r.columns: fig.add_trace(go.Scatter(x=r["date"], y=r[col], name=name, line=dict(color=color, width=2)), row=1, col=1)
    
    sm = summary.copy()
    for col in ["cagr", "vol", "max_dd", "hit_rate"]:
        if col in sm.columns: sm[col] = (sm[col] * 100).round(2).astype(str) + "%"
    for col in ["sharpe", "sortino", "calmar"]:
        if col in sm.columns: sm[col] = sm[col].round(2)
    fig.add_trace(go.Table(header=dict(values=list(sm.columns), fill_color="#1F4E78", font=dict(color="white", size=11), align="center"), cells=dict(values=[sm[c].tolist() for c in sm.columns], fill_color=[["#f7f6f2", "#ffffff"] * len(sm)], font=dict(color="#1a1a1a", size=11), align=["left"] + ["center"] * (len(sm.columns) - 1), height=25)), row=1, col=2)
    
    for col, name, color in [("dd_strategy", "Hybrid DD", COLORS["dd_hybrid"]), ("dd_buyhold", "B&H DD", COLORS["dd_bh"]), ("dd_logistic", "Logistic DD", COLORS["dd_lr"])]:
        if col in r.columns:
            fill_val = "tozeroy" if col == "dd_strategy" else "none"
            fill_color = "rgba(161, 44, 123, 0.2)" if col == "dd_strategy" else None
            fig.add_trace(go.Scatter(x=r["date"], y=r[col], name=name, line=dict(color=color, width=2 if col != "dd_strategy" else 1.5), fill=fill_val, fillcolor=fill_color), row=2, col=1)
    
    if "strategy_log_ret" in results.columns:
        strat_mo = results["strategy_log_ret"].resample("ME").sum().apply(lambda x: np.exp(x) - 1)
        bh_mo = results["buyhold_log_ret"].resample("ME").sum().apply(lambda x: np.exp(x) - 1)
        fig.add_trace(go.Histogram(x=strat_mo, name="Hybrid Monthly", nbinsx=40, marker_color=COLORS["hybrid"], opacity=0.7), row=2, col=2)
        fig.add_trace(go.Histogram(x=bh_mo, name="B&H Monthly", nbinsx=40, marker_color=COLORS["buyhold"], opacity=0.6), row=2, col=2)
    
    if "prob_up" in r.columns: fig.add_trace(go.Scatter(x=r["date"], y=r["prob_up"], name="NN Prob Up", line=dict(color=COLORS["prob_up"], width=1.5)), row=3, col=1)
    if "prob_best_regime" in r.columns: fig.add_trace(go.Scatter(x=r["date"], y=r["prob_best_regime"], name="HMM Best Regime", line=dict(color=COLORS["regime"], width=1.5, dash='dot')), row=3, col=1)
    if "signal" in r.columns: fig.add_trace(go.Bar(x=r["date"], y=r["signal"], name="Trade Signal", marker_color=COLORS["signal"], opacity=0.4), row=3, col=1)
    
    if "stable" in r.columns: fig.add_trace(go.Scatter(x=r["date"], y=r["stable"], name="Stability Mask", line=dict(color=COLORS["hybrid"], width=1), fill="tozeroy", fillcolor="rgba(1, 105, 111, 0.25)"), row=3, col=2)
    
    if len(hp_df) > 0 and "inner_sharpe" in hp_df.columns:
        fig.add_trace(go.Bar(x=hp_df["fold"].astype(str), y=hp_df["inner_sharpe"], name="Inner Sharpe", marker_color=COLORS["hybrid"]), row=4, col=1)
        if "n_states" in hp_df.columns: fig.add_trace(go.Scatter(x=hp_df["fold"].astype(str), y=hp_df["n_states"], name="n_states", mode="lines+markers", line=dict(color=COLORS["logistic"], width=3)), row=4, col=1)
    
    stat_data = [
        ["Sharpe (Point Est.)", f"{stat_tests.get('sharpe_point_estimate', np.nan):.3f}"],
        ["Sharpe 95% CI", f"[{stat_tests.get('sharpe_95_ci', (np.nan, np.nan))[0]:.3f}, {stat_tests.get('sharpe_95_ci', (np.nan, np.nan))[1]:.3f}]"],
        ["PSR vs Zero", f"{stat_tests.get('psr_vs_zero', np.nan):.1%}"],
        ["Deflated SR", f"{stat_tests.get('deflated_sharpe_ratio', np.nan):.1%}"],
        ["T-Statistic", f"{stat_tests.get('t_statistic', np.nan):.2f}"],
        ["P-Value", f"{stat_tests.get('t_pvalue', np.nan):.4f}"],
        ["Significant (5%)", "Yes" if stat_tests.get('significant_at_5pct', False) else "No"],
        ["Total Trials Tested", f"{stat_tests.get('n_trials_tested', 1)}"],
    ]
    fig.add_trace(go.Table(header=dict(values=["Metric", "Value"], fill_color="#2E7D32", font=dict(color="white", size=11), align="center"), cells=dict(values=[[row[0] for row in stat_data], [row[1] for row in stat_data]], fill_color=[["#f7f6f2", "#ffffff"] * len(stat_data)], font=dict(color="#1a1a1a", size=11), align=["left", "center"], height=22)), row=4, col=2)
    
    if "strategy_log_ret" in results.columns:
        rolling_sharpe = (results["strategy_log_ret"].rolling(63).mean() / results["strategy_log_ret"].rolling(63).std()) * np.sqrt(252)
        fig.add_trace(go.Scatter(x=results.index, y=rolling_sharpe, name="Rolling 63d Sharpe", line=dict(color=COLORS["hybrid"], width=2)), row=5, col=1)
        
        ann_ret = results["strategy_log_ret"].resample("YE").sum().apply(lambda x: np.exp(x) - 1)
        ann_bh = results["buyhold_log_ret"].resample("YE").sum().apply(lambda x: np.exp(x) - 1)
        years = ann_ret.index.year.astype(str)
        
        fig.add_trace(go.Bar(x=years, y=ann_ret, name="Hybrid Annual", marker_color=COLORS["hybrid"]), row=5, col=2)
        fig.add_trace(go.Bar(x=years, y=ann_bh, name="B&H Annual", marker_color=COLORS["buyhold"]), row=5, col=2)
        
        min_date, max_date = results.index[0], results.index[-1]
        fig.add_trace(go.Scatter(x=[min_date, max_date], y=[0, 0], mode="lines", line=dict(color="gray", dash="dash"), showlegend=False, hoverinfo="skip"), row=5, col=1)
        fig.add_trace(go.Scatter(x=[years[0], years[-1]], y=[0, 0], mode="lines", line=dict(color="gray", dash="dash"), showlegend=False, hoverinfo="skip"), row=5, col=2)
    
    fig.update_layout(template="plotly_white", height=2000, hovermode="x unified", legend=dict(orientation="h", yanchor="bottom", y=1.03, xanchor="left", x=0), margin=dict(l=60, r=40, t=60, b=50), barmode="group")
    fig.update_yaxes(title_text="Portfolio Value (Log)", row=1, col=1, type="log")
    fig.update_yaxes(title_text="Drawdown", tickformat=".1%", row=2, col=1)
    fig.update_yaxes(title_text="Rolling Sharpe", row=5, col=1)
    fig.update_yaxes(title_text="Annual Return", tickformat=".1%", row=5, col=2)
    return fig

def build_portfolio_figure(port_results: Dict[str, Any]):
    PCOLORS = ["#01696f", "#7a39bb", "#da7101", "#bab9b4", "#8d99ae", "#a12c7b", "#437a22", "#d19900"]
    fig = make_subplots(rows=3, cols=2, shared_xaxes=False, vertical_spacing=0.1, horizontal_spacing=0.06, row_heights=[0.4, 0.3, 0.3], column_widths=[0.6, 0.4], subplot_titles=("Portfolio vs Underlying Asset Equity Curves", "Portfolio Performance & Significance", "Portfolio Drawdown (Underwater)", "Asset Allocation Weights", "Monthly Return Distribution", "Asset Correlation Matrix"), specs=[[{"type": "xy"}, {"type": "table"}], [{"type": "xy"}, {"type": "pie"}], [{"type": "xy"}, {"type": "heatmap"}]])
    
    port_ret = port_results['portfolio_returns']
    port_eq = np.exp(port_ret.cumsum())
    fig.add_trace(go.Scatter(x=port_eq.index, y=port_eq, name="Hybrid Portfolio", line=dict(color="#1F4E78", width=3)), row=1, col=1)
    
    indiv = port_results['individual_results']
    c_idx = 0
    for ticker, data in indiv.items():
        t_ret = data['returns']
        t_eq = np.exp(t_ret.cumsum())
        fig.add_trace(go.Scatter(x=t_eq.index, y=t_eq, name=ticker, line=dict(color=PCOLORS[c_idx % len(PCOLORS)], width=1, dash='dot')), row=1, col=1)
        c_idx += 1
        
    stats = port_results['portfolio_stats']
    ["P-Value (vs 0)", f"{tests.get('t_pvalue', np.nan):.4f}"], ["Deflated Sharpe", f"{tests.get('deflated_sharpe_ratio', np.nan):.1%}"]
    fig.add_trace(go.Table(header=dict(values=["Metric", "Value"], fill_color="#1F4E78", font=dict(color="white", size=12), align="center"), cells=dict(values=[[row[0] for row in stat_data], [row[1] for row in stat_data]], fill_color=[["#f7f6f2", "#ffffff"] * len(stat_data)], font=dict(color="#1a1a1a", size=12), align=["left", "center"], height=25)), row=1, col=2)
    
    dd = port_eq / port_eq.cummax() - 1
    fig.add_trace(go.Scatter(x=dd.index, y=dd, name="Portfolio DD", line=dict(color="#a12c7b", width=1.5), fill="tozeroy", fillcolor="rgba(161, 44, 123, 0.2)"), row=2, col=1)
    
    weights = port_results['weights']
    fig.add_trace(go.Pie(labels=list(weights.keys()), values=list(weights.values()), hole=0.4, marker=dict(colors=PCOLORS)), row=2, col=2)
    
    port_mo = port_ret.resample("ME").sum().apply(lambda x: np.exp(x) - 1)
    fig.add_trace(go.Histogram(x=port_mo, name="Monthly Returns", nbinsx=40, marker_color="#01696f"), row=3, col=1)
    
    corr = port_results['correlation_matrix']
    fig.add_trace(go.Heatmap(z=corr.values, x=corr.columns, y=corr.columns, colorscale="RdBu", zmin=-1, zmax=1), row=3, col=2)
    
    fig.update_layout(template="plotly_white", height=1200, hovermode="x unified", margin=dict(l=60, r=40, t=60, b=50), showlegend=True)
    fig.update_yaxes(title_text="Cumulative Equity (Log)", type="log", row=1, col=1)
    fig.update_yaxes(title_text="Drawdown", tickformat=".1%", row=2, col=1)
    return fig

def generate_excel_bytes(results: pd.DataFrame, summary: pd.DataFrame, monthly: pd.DataFrame, hp_df: pd.DataFrame, stat_tests: Dict, config: Config) -> bytes:
    output = io.BytesIO()
    
    results.index.name = 'date'
    export = results.reset_index()
    monthly.index.name = 'month'
    month_exp = monthly.reset_index()
    
    stat_df = pd.DataFrame([{"metric": k, "value": v} for k, v in stat_tests.items() if not isinstance(v, (tuple, list, dict))])
    for key in ["sharpe_95_ci", "cagr_95_ci", "max_dd_95_ci", "sortino_95_ci"]:
        if key in stat_tests and stat_tests[key] is not None:
            ci = stat_tests[key]
            if isinstance(ci, tuple) and len(ci) == 2:
                stat_df = pd.concat([stat_df, pd.DataFrame([{"metric": f"{key}_lower", "value": ci[0]}, {"metric": f"{key}_upper", "value": ci[1]}])], ignore_index=True)
                
    config_dict = {f"{section}.{k}": v for section, section_data in asdict(config).items() for k, v in section_data.items()}
    config_df = pd.DataFrame(list(config_dict.items()), columns=["Parameter", "Value"])
    
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        export.to_excel(writer, sheet_name="EquityCurve", index=False)
        month_exp.to_excel(writer, sheet_name="MonthlyReturns", index=False)
        if not hp_df.empty: hp_df.to_excel(writer, sheet_name="HyperparamLog", index=False)
        stat_df.to_excel(writer, sheet_name="StatisticalTests", index=False)
        config_df.to_excel(writer, sheet_name="Configuration", index=False)
        
    wb = load_workbook(output)
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(name="Calibri", bold=True, color="FFFFFF")
    body_font = Font(name="Calibri", size=11, color="000000")
    pct_keywords = ['cagr', 'vol', 'max_dd', 'rate', 'prob', 'pct', 'return', 'ret', 'dd_', 'weight']
    
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 22
        if ws.dimensions != 'A1:A1': ws.auto_filter.ref = ws.dimensions
            
        headers = []
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            headers.append(str(cell.value).lower() if cell.value else "")
            
        for row in ws.iter_rows(min_row=2):
            for idx, cell in enumerate(row):
                cell.font = body_font
                cell.alignment = Alignment(vertical="center")
                if isinstance(cell.value, (float, int)):
                    col_name = headers[idx] if idx < len(headers) else ""
                    if any(kw in col_name for kw in pct_keywords) and 'sharpe' not in col_name and 'sortino' not in col_name and 'calmar' not in col_name:
                        cell.number_format = "0.00%"
                    else:
                        cell.number_format = "#,##0.0000"
                        
        for col in ws.columns:
            max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 26)

    if "EquityCurve" in wb.sheetnames:
        ws_eq = wb["EquityCurve"]
        chart = LineChart()
        chart.title = "Strategy Equity Curves"
        chart.style = 13
        chart.y_axis.title = 'Cumulative Equity'
        chart.x_axis.title = 'Date'
        headers = [str(cell.value).lower() if cell.value else "" for cell in ws_eq[1]]
        equity_cols = [i + 1 for i, h in enumerate(headers) if 'equity' in h]
        
        if equity_cols:
            data = Reference(ws_eq, min_col=min(equity_cols), max_col=max(equity_cols), min_row=1, max_row=ws_eq.max_row)
            cats = Reference(ws_eq, min_col=1, max_col=1, min_row=2, max_row=ws_eq.max_row)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.width = 30
            chart.height = 15
            dest_ws = wb["Summary"] if "Summary" in wb.sheetnames else ws_eq
            dest_ws.add_chart(chart, "A10")
            
    final_out = io.BytesIO()
    wb.save(final_out)
    return final_out.getvalue()

def generate_html_string(fig_html: str, title: str, stats: Dict) -> str:
    p_val = stats.get('t_pvalue', np.nan)
    sig_badge = "badge-green" if p_val < 0.05 else "badge-red"
    sig_text = "Statistically Significant" if p_val < 0.05 else "Not Significant"
    
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width,initial-scale=1">
  <title>{title} Dashboard</title>
  <style>
    body {{font-family: 'Inter', 'Helvetica Neue', Helvetica, Arial, sans-serif; margin: 0; background: #f4f4f4; color: #1a1a1a;}}
    .wrap {{max-width: 1600px; margin: 0 auto; padding: 40px 20px;}}
    .header-container {{background: #ffffff; padding: 30px; border-radius: 12px; box-shadow: 0 4px 12px rgba(0,0,0,0.05); margin-bottom: 24px;}}
    h1 {{margin: 0 0 12px; font-size: 34px; font-weight: 700; letter-spacing: -0.5px; color: #111;}}
    p {{color: #555; margin: 0 0 20px; font-size: 16px; line-height: 1.6;}}
    .card {{background: #fff; border: 1px solid #e2e2e2; border-radius: 12px; box-shadow: 0 8px 24px rgba(0,0,0,0.04); padding: 24px; overflow: hidden;}}
    .badge {{display: inline-block; padding: 5px 14px; border-radius: 6px; font-size: 13px; font-weight: 600; margin-right: 10px;}}
    .badge-green {{background: #d1fae5; color: #065f46;}}
    .badge-red {{background: #fee2e2; color: #991b1b;}}
    .badge-blue {{background: #dbeafe; color: #1e40af;}}
  </style>
</head>
<body>
<div class="wrap">
  <div class="header-container">
      <h1>{title} Hybrid Quantitative Dashboard</h1>
      <p>
        <span class="badge {sig_badge}">{sig_text} (p={p_val:.4f})</span>
        <span class="badge badge-blue">PSR: {stats.get('psr_vs_zero', np.nan):.1%}</span>
      </p>
      <p style="margin-bottom:0;">Production-grade Walk-Forward backtest with robust data fetching, sequence alignment, PyTorch NNs, and Cholesky HMMs.</p>
  </div>
  <div class="card">{fig_html}</div>
</div>
</body>
</html>"""

# ===========================================================================
# STREAMLIT UI
# ===========================================================================

def main():
    st.title("📈 Institutional Hybrid Markov + NN Backtester")
    st.markdown("End-to-end framework combining regime detection (HMM) with predictive Deep Learning (PyTorch).")
    
    mode = st.sidebar.radio("Backtest Mode", ["Single Asset", "Portfolio"])

    st.sidebar.header("Data Settings")
    if mode == "Single Asset":
        ticker = st.sidebar.text_input("Ticker Symbol", "SPY")
        tickers_list = [ticker]
    else:
        tickers_str = st.sidebar.text_area("Ticker Symbols (comma separated)", "SPY, QQQ, GLD, TLT")
        tickers_list = [t.strip() for t in tickers_str.split(",") if t.strip()]
        
    period = st.sidebar.selectbox("History Period", ["1y", "2y", "5y", "10y", "max"], index=3)
    interval = st.sidebar.selectbox("Interval", ["1d", "1wk"], index=0)

    st.sidebar.header("Model Architectures")
    architecture = st.sidebar.selectbox("NN Architecture", ["mlp", "resnet", "lstm", "transformer"])
    seq_len = st.sidebar.number_input("Sequence Length (LSTM/Transf)", value=20) if architecture in ["lstm", "transformer"] else 20
    n_states = st.sidebar.slider("HMM Hidden States", min_value=2, max_value=5, value=3)
    hidden_dim = st.sidebar.selectbox("Hidden Dimension", [32, 64, 128], index=1)
    use_student_t = st.sidebar.checkbox("Use Student-t HMM", value=False)
    use_mixed_precision = st.sidebar.checkbox("PyTorch AMP (GPU)", value=False)

    st.sidebar.header("Trading Constraints")
    prob_long = st.sidebar.slider("Probability Threshold (Long)", 0.50, 0.99, 0.52)
    prob_short = st.sidebar.slider("Probability Threshold (Short)", 0.01, 0.50, 0.48)
    regime_gate = st.sidebar.slider("Regime Confidence Gate", 0.0, 1.0, 0.45)
    vol_target = st.sidebar.number_input("Target Volatility", value=0.15)
    cost_bps = st.sidebar.number_input("Fixed Cost (bps)", value=2.0)
    spread_bps = st.sidebar.number_input("Spread (bps)", value=1.0)
    
    st.sidebar.header("Risk Management")
    enable_risk = st.sidebar.checkbox("Enable Stop Loss / Take Profit", value=False)
    stop_loss = st.sidebar.number_input("Stop Loss %", value=0.02, format="%.3f") if enable_risk else 0.02
    take_profit = st.sidebar.number_input("Take Profit %", value=0.05, format="%.3f") if enable_risk else 0.05
    max_dd_halt = st.sidebar.number_input("Max DD Halt %", value=0.10, format="%.2f")

    st.sidebar.header("Hyperparameter Tuning")
    enable_tune = st.sidebar.checkbox("Enable Optuna Tuning", value=True)
    n_trials = st.sidebar.slider("Optuna Trials per Fold", 1, 50, 5) if enable_tune else 0

    st.sidebar.header("Validation & Walk-Forward")
    n_splits = st.sidebar.slider("CV Splits", 2, 10, 5)
    anchored = st.sidebar.checkbox("Anchored Walk-Forward", value=False)

    if st.sidebar.button("🚀 Run Backtest", use_container_width=True, type="primary"):
        config = Config()
        config.data.period = period
        config.data.interval = interval
        config.model.architecture = architecture
        config.model.n_states = n_states
        config.model.hidden_dim = hidden_dim
        config.model.use_student_t_hmm = use_student_t
        config.model.seq_len = seq_len
        config.model.use_mixed_precision = use_mixed_precision
        config.trading.prob_long = prob_long
        config.trading.prob_short = prob_short
        config.trading.regime_gate = regime_gate
        config.trading.vol_target = vol_target
        config.trading.cost_bps = cost_bps
        config.trading.spread_bps = spread_bps
        config.risk.enabled = enable_risk
        config.risk.stop_loss_pct = stop_loss
        config.risk.take_profit_pct = take_profit
        config.risk.max_drawdown_halt = max_dd_halt
        config.tuning.enabled = enable_tune
        config.tuning.n_trials = n_trials
    config.backtest.n_splits = n_splits
    config.backtest.anchored = anchored

    if st.sidebar.button("🚀 Run Backtest", width="stretch", type="primary"):
        config = Config()
        config.data.period = period
        start_time = time.time()

        try:
            if mode == "Single Asset":
                config.data.ticker = tickers_list[0]
                results, summary, monthly, hp_df, stats = run_hybrid_backtest(config, progress_bar, status_text)
                fig = build_single_asset_figure(results, summary, hp_df, stats)
                
                # Store to session state for rendering
                st.session_state['mode'] = mode
                st.session_state['fig'] = fig
                st.session_state['results'] = results
                st.session_state['summary'] = summary
                st.session_state['monthly'] = monthly
                st.session_state['stats'] = stats
                
                # Build Downloads
                excel_bytes = generate_excel_bytes(results, summary, monthly, hp_df, stats, config)
                st.session_state['excel'] = excel_bytes
                st.session_state['html'] = generate_html_string(fig.to_html(full_html=False, include_plotlyjs="cdn"), f"{tickers_list[0]}", stats)
                
            else:
                port_res = run_portfolio_backtest(tickers_list, config, progress_bar, status_text)
                fig = build_portfolio_figure(port_res)
                
                st.session_state['mode'] = mode
                st.session_state['fig'] = fig
                st.session_state['port_res'] = port_res
                st.session_state['summary'] = pd.DataFrame([port_res['portfolio_stats']])
                
                # Build Downloads for Portfolio
                st.session_state['html'] = generate_html_string(fig.to_html(full_html=False, include_plotlyjs="cdn"), "Portfolio", port_res['portfolio_stat_tests'])

            status_text.success(f"Execution Completed in {time.time() - start_time:.1f}s")
            time.sleep(1)
            progress_bar.empty()
            status_text.empty()
            
        except Exception as e:
            st.error(f"Error during backtest: {str(e)}")
            progress_bar.empty()

    if 'fig' in st.session_state:
        tab1, tab2, tab3 = st.tabs(["📊 Dashboard View", "📋 Data & Metrics", "💾 Downloads"])
        
        with tab1:
            st.plotly_chart(st.session_state['fig'], width="stretch")
            
        with tab2:
            st.subheader("Performance Summary")
            st.dataframe(st.session_state['summary'], width="stretch")
            
            if st.session_state['mode'] == "Single Asset":
                col1, col2 = st.columns(2)
                with col1:
                    st.subheader("Statistical Tests")
                    stat_df = pd.DataFrame([{"Metric": k, "Value": v} for k, v in st.session_state['stats'].items() if not isinstance(v, (tuple, list, dict))])
                    # Fix PyArrow mixed-type serialization error
                    stat_df["Value"] = stat_df["Value"].astype(str)
                    st.dataframe(stat_df, width="stretch")
                with col2:
                    st.subheader("Monthly Returns")
                    st.dataframe(st.session_state['monthly'], width="stretch")
                    
                st.subheader("Raw Output (Tail)")
                st.dataframe(st.session_state['results'].tail(100), width="stretch")
            else:
                col1, col2 = st.columns(2)
                with col1:
                    st.subheader("Asset Weights")
                    st.dataframe(pd.DataFrame(list(st.session_state['port_res']['weights'].items()), columns=["Asset", "Weight"]))
                with col2:
                    st.subheader("Correlation Matrix")
                    st.dataframe(st.session_state['port_res']['correlation_matrix'])
                    
        with tab3:
            st.subheader("Export Artifacts")
            col1, col2, col3 = st.columns(3)
            with col1:
                if 'excel' in st.session_state:
                    st.download_button("📥 Download Master Excel", data=st.session_state['excel'], file_name="backtest_master.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                elif st.session_state['mode'] == 'Portfolio':
                    st.info("Portfolio Excel export not fully assembled in UI demo yet. CSVs available.")
            with col2:
                if 'html' in st.session_state:
                    st.download_button("📥 Download Interactive HTML", data=st.session_state['html'], file_name="dashboard.html", mime="text/html")
            with col3:
                if st.session_state['mode'] == "Single Asset":
                    csv = st.session_state['results'].to_csv()
                    st.download_button("📥 Download Raw Logs (CSV)", data=csv, file_name="raw_signals.csv", mime="text/csv")
                else:
                    csv = st.session_state['port_res']['portfolio_returns'].to_csv()
                    st.download_button("📥 Download Portfolio Returns", data=csv, file_name="portfolio_returns.csv", mime="text/csv")


if __name__ == "__main__":
    main()